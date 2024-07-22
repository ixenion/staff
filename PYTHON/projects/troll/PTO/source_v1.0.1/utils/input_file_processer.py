import asyncio
from calendar import TextCalendar
import openpyxl
# from openpyxl import Worksheet
import pandas as pd
import re

# from sys import path as sys_path
# sys_path.append("..")
# from ui.Terminal.terminal_control import Terminal


class InputFileProcesser:

    def __init__(self,
            # env_vars:dict,
            settings:object,
            calendar:object,
            uploads:object,
            progress_view: object,
            terminal: object,
            ) -> None:
        # self.env_vars = env_vars
        self.settings = settings
        self.calendar = calendar
        self.uploads = uploads
        self.progress_view = progress_view
        self.terminal = terminal

    async def start(self):
        """ MAIN of all main functions.
            It analyses all xlsx files (1C, TO1, TO2, exceptions);
            And creates & saves different output staff into output dir.
        """
        # Load input files
        file_1c, file_to1_opxl, file_to1, file_to2, file_exc = await self._load_input_files()

        # Gather all tbuses with odometer value
        # 'self.all_units' dict[unit_numb:str, odometer_value:float]
        # And sort them by odometer (odm) value
        all_units:dict[str, float] = await self._reaper_1c(file_1c) 
        # Send to terminal
        await self.terminal.log_add(f'Все машины:')
        for key in all_units:
            await self.terminal.log_add(f'{key}  |  {all_units[key]} км')
        await asyncio.sleep(0.1)
        # Update 'progress_view'
        self.progress_view.value += 0.04
        await self.progress_view.update_async()
        
        # Preprocess 'file_to1'
        file_to1_opxl_preprocessed, file_to1_meta \
                = await self._preprocess_to1(file_to1_opxl, file_to1)
        # Save (temporary) preprocessed (prepared) 'file_to1'
        # file_to1_opxl_preprocessed.save('123.xlsx')
        # Update 'progress_view'
        self.progress_view.value += 0.04
        await self.progress_view.update_async()

        # Extract exceptions from file_to2
        # 'self.exceptions' list[exceptional_unit_num:str]
        exceptions:list[str] = await self._reaper_to2(file_to2, exceptions=None)
        # Send to terminal
        await self.terminal.log_add(f'Машины-исключения:')
        for exc in exceptions:
            await self.terminal.log_add(f'{exc}')
        await asyncio.sleep(0.1)
        # Update 'progress_view'
        self.progress_view.value += 0.04
        await self.progress_view.update_async()

        # Extract exceptions from file_exc
        # 'self.exceptions' list[exceptional_unit_num:str]
        exceptions:list[str] = await self._reaper_exc(file_exc, exceptions=exceptions)
        # Send to terminal
        await self.terminal.log_add(f'Машины-исключения:')
        for exc in exceptions:
            await self.terminal.log_add(f'{exc}')
        await asyncio.sleep(0.1)
        # Update 'progress_view'
        self.progress_view.value += 0.04
        await self.progress_view.update_async()

        return all_units, exceptions, file_to1_opxl_preprocessed, file_to1_meta


    async def _reaper_exc(self, file_exc:pd.DataFrame, exceptions:list[str]|None) -> list[str]:
        """ Function exstracts tbus units' numbers from exception.xlsx
            And adds them to the exceptions list."""
        await self.terminal.log_add(f'\n[*] Анализ файла ИСКЛ ...\n')

        async def get_unit_number_begin_coords(file_exc:pd.DataFrame) -> list[int]:
            """ Analyze input file to locate (row, column)
                where units number are started."""
            file_not_empty = False
            # Init pattern (4digits 4nonDigits)
            # pattern = re.compile(r'^\d{4}\s\D{4}$', re.I)
            pattern = re.compile(r'^\d{4}', re.I)
            # Get file 'dimentions' (rows-height and columns-width) to proper iteration
            # while finding units and odometer columns.
            rows, columns = file_exc.shape
            if rows == 0 or columns == 0:
                return [-1, -1]

            for row in range(rows):
                for column in range(columns):
                    cell = str(file_exc.loc[row][column])
                    if cell:
                        file_not_empty = True
                    matches = pattern.findall(cell)
                    for _ in matches:
                        return [row, column]
            if file_not_empty:
                await self.terminal.log_add(f'[-] ОШИБКА: Невозможно определить столбец с номерами исключений.\t\nНомер машин должен быть целым четырехзначным числом.')
                raise RuntimeError

        async def unit_num_check_value(unit_num:str) -> str:
            """ Extract unit number,
                If unsuccessful - error raised."""
            # if not unit_num and len(all_units) == 0:
            #     return unit_num
            # Cant use this pattern - for some reasons its not working...
            # pattern = re.compile(r'^\d{4}\s[(тб)]$', re.I)
            if not unit_num:
                return unit_num
            pattern = re.compile(r'^\d{4}', re.I)
            matches = pattern.findall(unit_num)
            for match in matches:
                unit_num = match[:4]
                return unit_num
            await self.terminal.log_add(f'[-] ОШИБКА: посторонний обьект в столбце с номерами машин "unit_number".\n\tСтолбец номеров машин должен содержать значения типа "####", например "1001" или "1001 (тб).')
            raise RuntimeError

        if exceptions is None:
            exceptions = []
        if file_exc is None:
            # await self.terminal.log_add(f'[✔] Анализ файла ИСКЛ завершен успешно.\n')
            return exceptions
        
        unit_num_start_row, unit_num_start_column = await get_unit_number_begin_coords(file_exc)
        if unit_num_start_row == -1:
            return exceptions
        # Get shape of file_exc
        file_exc_rows, file_exc_columns = file_exc.shape
        for row in range(unit_num_start_row, file_exc_rows):
            unit_exc = str(file_exc.loc[row][unit_num_start_column])
            unit_exc = await unit_num_check_value(unit_exc)
            if not unit_exc:
                continue
            exceptions.append(unit_exc)
        # Remove duplicates
        exceptions = list(dict.fromkeys(exceptions))
        await self.terminal.log_add(f'[+] Анализ файла ИСКЛ завершен успешно.\n')
        return exceptions

    async def _reaper_to2(self, file_to2:pd.DataFrame, exceptions:list[str]|None) -> list[str]:
        """ day_control value used here.
            all tbus units after that day value (including this day)
            Are excluded (added to exceptions).
        """
        await self.terminal.log_add(f'\n[*] Анализ файла TO2 ...\n')

        async def get_unit_number_begin_coords(file_to2:pd.DataFrame) -> list[int]:
            """ Analyze input file to locate (row, column)
                where units number are started."""
            rows, columns = file_to2.shape
            for row in range(rows):
                for column in range(columns):
                    cell = str(file_to2.loc[row][column])
                    if cell == 'Число':
                        day = str(self.settings.day_control.value)
                        for day_row in range(row, rows):
                            cell = file_to2.loc[day_row][column]
                            if cell == day:
                                return [day_row, column]
                        await self.terminal.log_add(f'[-] ОШИБКА: в файле ТО2 отсутствует Число "{self.settings.day_control.value}".')
                        raise RuntimeError
            await self.terminal.log_add(f'[-] ОШИБКА: Невозможно определить столбец с числами.\t\nСтолбец должен иметь название "Число".')
            raise RuntimeError

        async def unit_num_check_value(unit_num:str) -> str:
            """ Extract unit number,
                If unsuccessful - skip."""
            # if not unit_num and len(all_units) == 0:
            #     return unit_num
            # Cant use this pattern - for some reasons its not working...
            # (because need to use r'^\d{4}\s(тб)' ???)
            # pattern = re.compile(r'^\d{4}\s[(тб)]$', re.I)
            if not unit_num:
                return unit_num
            pattern = re.compile(r'^\d{4}', re.I)
            matches = pattern.findall(unit_num)
            for match in matches:
                unit_num = match[:4]
                return unit_num
            # await self.terminal.log_add(f'[✘] ОШИБКА: посторонний обьект в столбце с номерами машин "unit_number".\n\tСтолбец номеров машин должен содержать значения типа "####", например "1001" или "1001 (тб).')
            # raise RuntimeError

        if exceptions is None:
            exceptions = []

        day_row, day_column = await get_unit_number_begin_coords(file_to2)
        rows, columns = file_to2.shape
        for row in range(day_row, rows):
            for column in range(columns):
                cell = str(file_to2.loc[row][column])
                unit_num = await unit_num_check_value(cell)
                if unit_num:
                    exceptions.append(unit_num)
                pass

        # Remove duplicates
        exceptions = list(dict.fromkeys(exceptions))
        await self.terminal.log_add(f'[+] Анализ файла TO2 завершен успешно.\n')
        return exceptions

    async def _preprocess_to1(self, file_to1_opxl:pd.DataFrame, file_to1):
        """ Function takes TO1 file as input,
            Fills 2 days rows with next month values,
            Clears all cells to the right from tbus units' numbers columns.
        """

        await self.terminal.log_add(f'\n[*] Подготовка шаблона файла TO1 ...\n')
        self.file_to1_rows, self.file_to1_columns = file_to1.shape

        first_sheet_name = file_to1_opxl.sheetnames[0]
        file_to1_opxl_first_sheet = file_to1_opxl[first_sheet_name]

        async def get_days_top_left_cell_coord() -> tuple[int, int]:
            """ Find start Row Column of double rowed days field."""
            pattern = re.compile(r'^бригада', re.I)
            # sheet_rows = file_to1.max_row
            # Too many columns for some reason... Cant rely on this param
            # sheet_columns = file_to1.max_column
            # Openpyxl starts intexing from '1', not '0'
            for row in range(1,self.file_to1_rows+1):
                for column in range(1,self.file_to1_columns+1):
                    cell = str(file_to1_opxl_first_sheet.cell(row,column).value)
                    if cell is None:
                        continue
                    matches = pattern.findall(cell)
                    for _ in matches:
                        # Found coords of the first 'бригада*' cell.
                        # That means - start Row Column of
                        # double rowed days field would be row+1, column.
                        day_row = row+1
                        # Check that this cell really contains day data.
                        pattern = re.compile(r'^\d{1,2}$', re.I)
                        day_cell = str(file_to1_opxl_first_sheet.cell(day_row,column).value)
                        matches = pattern.findall(day_cell)
                        for _ in matches:
                            return day_row, column
                        await self.terminal.log_add(f'\n[-] ОШИБКА: Неверный формат даты: "{day_cell}" в ячейке {day_row}:{column}.')
                        raise RuntimeError
            await self.terminal.log_add(f'\n[-] ОШИБКА: Невозможно найти ячейку со значением "бригада №".')
            raise RuntimeError

        async def fill_days_rows_next_month(start_row:int, start_column:int) -> None:
            """ Fill days' rows with next month days. """
            async def is_correct(day:str) -> int:
                pattern = re.compile(r'^\d{1,2}$', re.I)
                matches = pattern.findall(day)
                for _ in matches:
                    return int(day)
                await self.terminal.log_add(f'\n[-] ОШИБКА: Неверный формат даты: "{day}" в ячейке {row}:{column}.')
                raise RuntimeError
                return 1
                
            # Find fill_start coords
            # Fill row starts always in one particular row
            fill_start_row = start_row
            # +2 in case top left is max value
            fill_start_column = start_column+2
            last_day_value = int(file_to1_opxl_first_sheet.cell(start_row, start_column).value)
            # Checking only 2 rows
            for row in range(start_row, start_row+2):
                # But all columns
                for column in range(start_column, self.file_to1_columns+1):
                    cell = file_to1_opxl_first_sheet.cell(row, column).value
                    # Skip empty cells
                    if cell is None:
                        continue
                    cell:int = await is_correct(str(cell))
                    if cell > last_day_value:
                        last_day_value = cell
                        if (column - start_column)%2 == 0:
                            fill_start_column = column+2
                        else:
                            fill_start_column = column+1
            # print(fill_start_row, fill_start_column)
            # Filling algorithm
            cal = TextCalendar()
            year = int(self.settings.year_control.value)
            month = int(self.settings.month_control.value)
            month_cal = cal.monthdays2calendar(year, month)
            month_days = []
            for week in month_cal:
                for day in week:
                    if day[0]:
                        month_days.append(day[0])
            
            #----------------|
            #----------------|
            for row in range(start_row, start_row+2):
                for column in range(start_column, self.file_to1_columns+1):
                    file_to1_opxl_first_sheet.cell(row, column).value = None

            #        + + + + |
            #                |
            for column in range(fill_start_column, self.file_to1_columns+1, 2):
                file_to1_opxl_first_sheet.cell(fill_start_row, column).value = month_days[0]
                month_days.pop(0)
            #+ + + + + + + + |
            #                |
            for column in range(start_column, self.file_to1_columns+1, 2):
                if file_to1_opxl_first_sheet.cell(fill_start_row, column).value is None:
                    file_to1_opxl_first_sheet.cell(fill_start_row, column).value = month_days[0]
                    month_days.pop(0)
                else:
                    break
            #+ + + + ++++++++|
            #                |
            for column in range(fill_start_column+1, self.file_to1_columns+1, 2):
                file_to1_opxl_first_sheet.cell(fill_start_row, column).value = month_days[0]
                month_days.pop(0)
            #++++++++++++++++|
            #                |
            for column in range(start_column+1, self.file_to1_columns+1, 2):
                if file_to1_opxl_first_sheet.cell(fill_start_row, column).value is None:
                    file_to1_opxl_first_sheet.cell(fill_start_row, column).value = month_days[0]
                    month_days.pop(0)
                else:
                    break
            #++++++++++++++++|
            #        + + + + |
            for column in range(fill_start_column, self.file_to1_columns+1, 2):
                file_to1_opxl_first_sheet.cell(fill_start_row+1, column).value = month_days[0]
                month_days.pop(0)
            #++++++++++++++++|
            #+ + + + + + + + |
            for column in range(start_column, self.file_to1_columns+1, 2):
                if file_to1_opxl_first_sheet.cell(fill_start_row+1, column).value is None:
                    file_to1_opxl_first_sheet.cell(fill_start_row+1, column).value = month_days[0]
                    month_days.pop(0)
                else:
                    break
            #++++++++++++++++|
            #+ + + + ++++++++|
            for column in range(fill_start_column+1, self.file_to1_columns+1, 2):
                if len(month_days) > 0:
                    file_to1_opxl_first_sheet.cell(fill_start_row+1, column).value = month_days[0]
                    month_days.pop(0)
            #++++++++++++++++|
            #++++++++++++++++|
            for column in range(start_column+1, self.file_to1_columns+1, 2):
                if file_to1_opxl_first_sheet.cell(fill_start_row+1, column).value is None \
                        and len(month_days) > 0:
                    file_to1_opxl_first_sheet.cell(fill_start_row+1, column).value = month_days[0]
                    month_days.pop(0)
                else:
                    break

        async def clear_units_status_cell(start_row:int, start_column:int) -> None:
            """ Clear cells to the right of units' numbers columns."""
            pattern = re.compile(r'^\d{4}$', re.I)
            for row in range(start_row+2, self.file_to1_rows+1):
                for column in range(start_column, self.file_to1_columns+1, 2):
                    # Check is that unit number, if no - skip.
                    cell = str(file_to1_opxl_first_sheet.cell(row, column).value)
                    matches = pattern.findall(cell)
                    for _ in matches:
                        file_to1_opxl_first_sheet.cell(row, column+1).value = None

        async def change_data_header() -> None:
            pattern = re.compile(r'^на', re.I)
            for row in range(1, self.file_to1_rows+1):
                for column in range(1, self.file_to1_columns+1):
                    cell = str(file_to1_opxl_first_sheet.cell(row, column).value)
                    matches = pattern.findall(cell)
                    for _ in matches:
                        file_to1_opxl_first_sheet.cell(row, column).value = f'на {self.calendar.date_view.value}'
            
        
        # Get days top left cell coords
        top_left_day_row, top_left_day_column = await get_days_top_left_cell_coord()
        # Fill days' rows with next month days
        await fill_days_rows_next_month(top_left_day_row, top_left_day_column)
        # Clear cells to the right of units' numbers columns
        await clear_units_status_cell(top_left_day_row, top_left_day_column)
        # Change DATA in the header
        await change_data_header()

        file_to1_opxl_preprocessed = file_to1_opxl
        file_to1_opxl_meta = [
                top_left_day_row,
                top_left_day_column,
                self.file_to1_rows,
                self.file_to1_columns,]

        await self.terminal.log_add(f'[+] Подготовка шаблона файла TO1 завершена успешно.\n')
        return file_to1_opxl_preprocessed, file_to1_opxl_meta

    # REAPER 1C
    async def _reaper_1c(self, file_1c:pd.DataFrame) -> dict[str, float]:
        """ Purpose of this function is to extract tbus numbers (called units),
            And corresponding odometer values.
        """
        
        await self.terminal.log_add(f'[*] Анализ файла 1С ...\n')

        async def get_unit_number_begin_coords(file_1c:pd.DataFrame) -> list[int]:
            """ Analyze input file to locate (row, column)
                where units number are started."""
            # Init pattern (4digits 4nonDigits)
            # pattern = re.compile(r'^\d{4}\s\D{4}$', re.I)
            pattern = re.compile(r'^ТС$', re.I)
            # Get file 'dimentions' (rows-height and columns-width) to proper iteration
            # while finding units and odometer columns.
            rows, columns = file_1c.shape
            for row in range(rows):
                for column in range(columns):
                    cell = str(file_1c.loc[row][column])
                    matches = pattern.findall(cell)
                    for _ in matches:
                        # +2 because 'TC' takes 2 rows
                        return [row+1, column]
            await self.terminal.log_add(f'[-] ОШИБКА: Невозможно определить столбец с номерами троллейбусов в файле 1С.xlsx\n\tСтолбец номеров машин должен иметь название "ТС".')
            raise RuntimeError

        async def get_unit_odometer_begin_coords(file_1c:pd.DataFrame) -> list[int]:
            """ Analyze input file to locate (row, column)
                where units odometer are started."""
            # Init pattern ()
            # pattern = re.compile(r'^\D{6}\s\D{5}\s\D{3}[2]$', re.I)
            pattern = re.compile(r'^Пробег после ТО-2$', re.I)
            # Get file 'dimentions' (rows-height and columns-width) to proper iteration
            # while finding units and odometer columns.
            rows, columns = file_1c.shape
            for row in range(rows):
                for column in range(columns):
                    cell = str(file_1c.loc[row][column])
                    matches = pattern.findall(cell)
                    for _ in matches:
                        return [row+1, column]
            await self.terminal.log_add(f'[-] ОШИБКА: Невозможно определить столбец с пробегом троллейбусов в файле 1С.xlsx\n\tСтолбец пробега должен иметь название "Пробег после ТО-2".')
            raise RuntimeError
        
        async def unit_num_check_value(unit_num:str) -> str:
            """ Extract unit number,
                If unsuccessful - error raised."""
            # if not unit_num and len(all_units) == 0:
            #     return unit_num
            # Cant use this pattern - for some reasons its not working...
            # pattern = re.compile(r'^\d{4}\s[(тб)]$', re.I)
            pattern = re.compile(r'^\d{4}', re.I)
            matches = pattern.findall(unit_num)
            for match in matches:
                unit_num = match[:4]
                return unit_num
            await self.terminal.log_add(f'[-] ОШИБКА: Неверный формат номера машины "{unit_num}".\n\tСтолбец номеров машин должен содержать значения типа "####", например "1001" или "1001 (тб)".\n\tТак же убедитесь, что с последним номером машини заканчивается вся таблица.')
            raise RuntimeError

        async def unit_odm_check_value(unit_odm:float) -> float:
            if unit_odm == '':
                return 0.0
            try:
                unit_odm = float(unit_odm)
                return unit_odm
            except ValueError:
                await self.terminal.log_add(f'[-] ОШИБКА: Неверный формат значения в столбце пробега: "{unit_odm}".\n\tДолжно быть не отрицательное рациональное число.')
                raise RuntimeError

        # Init result variable
        all_units = {}
        # Get units number column start position 
        unit_num_start_row, unit_num_start_column = await get_unit_number_begin_coords(file_1c)
        unit_odm_start_row, unit_odm_start_column = await get_unit_odometer_begin_coords(file_1c)
        # Get file_1c shape
        file_1c_rows, file_1c_columns = file_1c.shape
        # multiple rows in 'TC' issue
        start_row = max(unit_num_start_row, unit_odm_start_row)
        # Gather 'unit_num_value' and 'unit_odm_value' together into dict[str, float]
        for row in range(start_row, file_1c_rows):
            unit_num_value = file_1c.loc[row][unit_num_start_column]
            #  Get rid of ' (тб)' and check that 'unit_num_value' is correct.
            unit_num_value = await unit_num_check_value(unit_num_value)
            # Ensure that odometer value is correct and float
            unit_odm_value = file_1c.loc[row][unit_odm_start_column]
            unit_odm_value = await unit_odm_check_value(unit_odm_value)
            # all_units[unit_num_value] = unit_odm_value
            if unit_num_value not in all_units:
                all_units[unit_num_value] = unit_odm_value
            else:
                await self.terminal.log_add(f'[-] ОШИБКА: Дублируется машина "{unit_num_value}".')
                raise RuntimeError
        # Sort dict by value (odometer)
        all_units = dict(sorted(all_units.items(), key=lambda item: item[1], reverse=True))

        await self.terminal.log_add(f'[+] Анализ файла 1С завершен успешно.\n')
        return all_units

    async def _load_input_files(self) -> list[pd.DataFrame]:
        """ Loads all four files with help of pandas library. """
        # FILE 1C
        path_1c = self.uploads.fpUnit1.path
        try:
            file_1c = pd.read_excel(path_1c,na_filter=False,header=None,)
        except Exception as e:
            await self.terminal.log_add(f'[-] Неверный путь к файлу 1С.\n')
            raise e
        # FILE TO1
        path_to1 = self.uploads.fpUnit2.path
        try:
            # file_to1 = pd.read_excel(path_to1, na_filter=False, header=None)
            file_to1_opxl = openpyxl.load_workbook(path_to1, data_only=True)
            # first_sheet_name = file_to1.sheetnames[0]
            # file_to1_opxl_first_sheet = file_to1[first_sheet_name]
            file_to1 = pd.read_excel(self.uploads.fpUnit2.path, na_filter=False, header=None)
        except Exception as e:
            await self.terminal.log_add(f'[-] Неверный путь к файлу TO1.\n')
            raise e
        # FILE TO2
        path_to2 = self.uploads.fpUnit3.path
        try:
            file_to2 = pd.read_excel(path_to2, na_filter=False, header=None)
        except Exception as e:
            await self.terminal.log_add(f'[-] Неверный путь к файлу TO2.\n')
            raise e
        # FILE EXCEPTIONS
        path_exc = self.uploads.fpUnit4.path
        file_exc = None
        try:
            if path_exc:
                file_exc = pd.read_excel(path_exc, na_filter=False, header=None)
        except Exception as e:
            await self.terminal.log_add(f'[-] Неверный путь к файлу исключений.\n')
            raise e
        # result = [file_1c, file_to1, file_to2, file_exc]
        result = [file_1c, file_to1_opxl, file_to1, file_to2, file_exc]
        return result

    async def start_test_payload(self) -> None:
        """ Test function, to simulate work of 'self.start' function."""
        for unit in range(1000, 1050):
            await self.terminal.log_add(unit)
            self.progress_view.value += 0.02
            await self.progress_view.update_async()
            await asyncio.sleep(0.1)
            # payload stress test
            # if unit == 1025:
            #     1/0
        pass
