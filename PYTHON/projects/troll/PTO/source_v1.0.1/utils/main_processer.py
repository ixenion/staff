from math import ceil
from os import path as os_path
from openpyxl import Workbook
import re
from typing import Any


class MainProcesser:

    def __init__(
            self,
            env_vars:dict[str,Any],
            uploads:object,
            progress_view:object,
            terminal:object,
            all_units:dict[str,float],
            exceptions:list[str],
            file_to1,
            file_to1_meta,
            ):
        self.env_vars = env_vars
        self.uploads = uploads
        self.progress_view = progress_view
        self.terminal = terminal
        
        self.all_units = all_units
        self.exceptions = exceptions
        self.file_to1 = file_to1
        self.file_to1_meta = file_to1_meta

    async def start(self):
        """ Start main composing algorithm."""

        # Unpack all settings variables
        electro_tbus_list = self.env_vars['pto_electro_tbus_list']
        lengthy_tbus_list = self.env_vars['pto_lengthy_tbus_list']
        working_days_list = self.env_vars['working_days_list']
        non_working_days_list = self.env_vars['non_working_days_list']

        max_etbus_per_day = int(self.env_vars['pto_etbus_control_value'])
        max_ltbus_per_day = int(self.env_vars['pto_ltbus_control_value'])
        max_tbus_per_lday = float(self.env_vars['pto_all_ltbus_control_value'])
        max_tbus_per_day = float(self.env_vars['pto_alltbus_control_value'])
        
        min_odometer_value = float(self.env_vars['pto_min_odometer_value'])
        save_dir_path = self.env_vars['pto_dir1_path']

        ######################################################################
        # Generate rest variables

        oscilating_day_unit_nimber = False if max_tbus_per_day%1 == 0 else True
        max_tbus_per_day = ceil(max_tbus_per_day)
        oscilating_lday_unit_nimber = False if max_tbus_per_lday%1 == 0 else True
        max_tbus_per_lday = ceil(max_tbus_per_lday)

        # Find what days can be used by every single unit
        # KEY unit:str  VALUE [days:int]
        # KEY '1001'    VALUE [3, 11, 19, 27]
        all_units_possible_days = {}

        top_left_days_row       = self.file_to1_meta[0]
        top_left_days_column    = self.file_to1_meta[1]
        file_to1_rows           = self.file_to1_meta[2]
        file_to1_columns        = self.file_to1_meta[3]
        file_to1_first_sheet_name = self.file_to1.sheetnames[0]
        file_to1_first_sheet = self.file_to1[file_to1_first_sheet_name]
        pattern = re.compile(r'^\d{4}', re.I)
        for row in range(top_left_days_row+2, file_to1_rows+1):
            for column in range(top_left_days_column, file_to1_columns+1, 2):
                unit = str(file_to1_first_sheet.cell(row, column).value)
                matches = pattern.findall(unit)
                for _ in matches:
                    all_units_possible_days[unit] = [
                            file_to1_first_sheet.cell(top_left_days_row, column).value,
                            file_to1_first_sheet.cell(top_left_days_row, column+1).value,
                            file_to1_first_sheet.cell(top_left_days_row+1, column).value,
                            file_to1_first_sheet.cell(top_left_days_row+1, column+1).value,
                            ]
                    # Remove all 'None' values
                    all_units_possible_days[unit] = \
                            list(filter(lambda elem: elem is not None,
                                all_units_possible_days[unit]))

        # KEY day:int   VALUE [units:str]
        # KEY 14        VALUE ['1001', '1002', '1003', '1004', '1005']
        raw_output = {}

        # KEY unit:str  VALUE [day:int, odm:float]
        # KEY '1001'    VALUE [14, 8234.6]
        raspisanie = {}
        
        # KEY unit:str  VALUE [odm:float]
        # KEY '1001'    VALUE [8234.6]
        ostatok = self.all_units.copy()
        # Remove 'exceptions' from 'ostatok'
        for unit in self.exceptions:
            if unit in ostatok:
                del ostatok[unit]

        ######################################################################
        # Functions

        async def day_have_free_space(day:int, unit:str) -> bool:

            if day not in working_days_list:
                # raw_output[day] = []
                return False

            if day not in raw_output:
                raw_output[day] = []
                return True
            
            # Get max unit per day
            if not await day_have_ltbus(day):
                max_tbus_per_day_local = max_tbus_per_day
                # Oscilation happens here
                if oscilating_lday_unit_nimber:
                    # If day umber: even - osc UP
                    if day%2==0:
                        pass
                    else:
                        max_tbus_per_day_local = max_tbus_per_day_local - 1
            else:
                max_tbus_per_day_local = max_tbus_per_lday
                # Oscilation happens here
                if oscilating_day_unit_nimber:
                    # If day umber: even - osc UP
                    if day%2==0:
                        pass
                    else:
                        max_tbus_per_day_local = max_tbus_per_day_local - 1

            # Accign unit
            if await is_ltbus(unit):
                # Check how much ltbus we already have in this day
                if not await day_have_ltbus(day) < max_ltbus_per_day:
                    return False
                if len(raw_output[day]) < max_tbus_per_day_local:
                    return True
                return False

            elif await is_etbus(unit):
                # Check how much etbus we already have in this day
                if not await day_have_etbus(day) < max_etbus_per_day:
                    return False
                if len(raw_output[day]) < max_tbus_per_day_local:
                    return True
                return False

            else:
                if len(raw_output[day]) < max_tbus_per_day_local:
                    return True
                return False

        async def is_etbus(unit:str) -> bool:
            for pattern in electro_tbus_list:
                if await match_pattern(unit, pattern):
                    return True
            return False

        async def is_ltbus(unit:str) -> bool:
            for pattern in lengthy_tbus_list:
                if await match_pattern(unit, pattern):
                    return True
            return False

        async def match_pattern(unit:str, pattern:str) -> bool:
            """ Check that pattern and unit have equal length."""
            if not len(unit) == len(pattern):
                await self.terminal.log_add(f'[-] Входное значение {unit} не соответствует паттерну {pattern}.')
                raise RuntimeError
            for i, element in enumerate(pattern):
                if element == '#':
                    continue
                if unit[i] == element:
                    return True
            return False
        
        async def day_have_etbus(day:int) -> int:
            count = 0
            for unit in raw_output[day]:
                for pattern in electro_tbus_list:
                    if await match_pattern(unit, pattern):
                        count += 1
            return count

        async def day_have_ltbus(day:int) -> int:
            count = 0
            for unit in raw_output[day]:
                for pattern in lengthy_tbus_list:
                    if await match_pattern(unit, pattern):
                        count += 1
            return count

        def save_output_files() -> None:
            save_path = save_dir_path
            raw_output_file_name    = 'raw_output.xlsx'
            raspisanie_file_name    = 'raspisanie.xlsx'
            ostatok_file_name       = 'ostatok.xlsx'
            file_to1_file_name      = 'TO1.xlsx'
            path_raw_output = os_path.join(save_path, raw_output_file_name)
            path_raspisanie = os_path.join(save_path, raspisanie_file_name)
            path_ostatok = os_path.join(save_path, ostatok_file_name)
            path_file_to1 = os_path.join(save_path, file_to1_file_name)
            # Save 'raw_output'
            # df = pd.DataFrame(list(zip( list(raw_output.keys()), raw_output.values() )))
            # df.to_excel(path_raw_output)
            wb = Workbook()
            ws = wb.active
            ws.append(['День', 'Номера машин'])
            for day in list(raw_output.keys()):
                ws.append([day, *raw_output[day]])
            wb.save(path_raw_output)
            # Save 'raspisanie'
            wb = Workbook()
            ws = wb.active
            ws.append(['Номер', 'День', 'Пробег'])
            for unit in list(raspisanie.keys()):
                ws.append([unit, *raspisanie[unit]])
            wb.save(path_raspisanie)
            # Save 'ostatok'
            wb = Workbook()
            ws = wb.active
            ws.append(['Номер', 'Пробег', 'Возможные дни'])
            for unit in list(ostatok.keys()):
                ws.append([unit, *ostatok[unit]])
            wb.save(path_ostatok)
            # Save NEW TO1
            self.file_to1.save(path_file_to1)

        def postprocess_raw_output(raw_output:dict[int, list[str]]) -> dict[int, list[str]]:
            """ There is a broblem with 'raw_otput' - if day has no any unit it just skipped
                Example (days):
                    1 1001, 1002
                    2 1003, 1004
                    5 1005, 1006
                    6 1007, 1008
                    (See? there are no numbers 3 and 4)
                So purpose of this function is to correct this like so:
                    1 1001, 1002
                    2 1003, 1004
                    3
                    4
                    5 1005, 1006
                    6 1007, 1008
            """
            
            # Get all month days
            all_days = [*working_days_list, *non_working_days_list]

            # Add missing (empty) days
            for day in all_days:
                if day in raw_output:
                    # That means 'raw_output' already has data in that day
                    pass
                else:
                    raw_output[day] = []

            # Sort 'raw_output' by days (key)
            raw_output = dict(sorted(raw_output.items()))

            return raw_output

        # MAIN sceduler
        # How much to add to the progress bar
        progress_add = 1.0 / len(self.all_units)
        # form 'raw_output'
        # and 'ostatok'
        # and 'raspisanie'
        await self.terminal.log_add(f'\n[*] Составление нового графика ...')
        for unit in self.all_units:
            
            # Skip exception units
            if unit in self.exceptions:
                continue
            
            unit_odm = self.all_units[unit]
            # Skip units with too small odometer value
            if unit_odm < min_odometer_value:
                continue
            
            # There are may be units in 1C which are not presented at TO1 file.
            # So - raise exception
            try:
                possible_days = all_units_possible_days[unit]
            # except Exception as e:
            except KeyError:
                await self.terminal.log_add(f'\n[-] ОШИБКА: В файле ТО1 отсутствует машина "{unit}", которая есть в файле 1С.')
                raise RuntimeError

            # Find suitable day for unit
            for possible_day in possible_days:
                # Check that day has enought space for this unit
                if await day_have_free_space(possible_day, unit):
                    # Update 'raw_output' (Add unit to this day)
                    raw_output[possible_day].append(unit)
                    # Update 'raspisanie' (add unit,day,odm)
                    raspisanie[unit] = ([possible_day, unit_odm])
                    # Update 'ostatok' (Remove this unit from 'ostatoc')
                    del ostatok[unit]
                    # Update TO1.xlsx
                    for row in range(top_left_days_row+2, file_to1_rows):
                        for column in range(top_left_days_column, file_to1_columns):
                            cell = str(file_to1_first_sheet.cell(row, column).value)
                            if unit == cell:
                                file_to1_first_sheet.cell(row, column+1).value = possible_day
                    # Progress ring
                    self.progress_view.value += progress_add
                    await self.progress_view.update_async()
                    # sleep(0.1)
                    break
        
        # For debug purpose
        # Sort 'raw_output' by days (key)
        # raw_output = dict(sorted(raw_output.items()))

        # Add missing days
        # and sort 'raw_output' by days (key)
        raw_output = postprocess_raw_output(raw_output)

        # Sort 'raspisanie' by units (key)
        raspisanie = dict(sorted(raspisanie.items()))
        
        # 'ostatok' already sorted by odometer
        # Add possible days into 'ostatok'
        # KEY unit:str  VALUE [odm:float, [team_days]]
        # KEY '1001'    VALUE [8234.6, [3, 11, 19, 27]]
        for unit in ostatok:
            possible_days = all_units_possible_days[unit]
            ostatok[unit] = [ostatok[unit], *possible_days]
        
        # Save all data
        save_output_files()

        # STDOUT
        # print(f'\nraw_output')
        # for key in raw_output:
        #     print(f'{key} {raw_output[key]}')
        
        # print(f'\nraspisanie')
        # for key in raspisanie:
        #     print(f'{key} {raspisanie[key]}')

        # print(f'\nostatok')
        # for key in ostatok:
        #     print(f'{key} {ostatok[key]}')

        # Clear memory
        del all_units_possible_days
        del raw_output
        del raspisanie
        del ostatok
        del self.file_to1

        await self.terminal.log_add(f'\n[+] Составление нового графика завершено.')
