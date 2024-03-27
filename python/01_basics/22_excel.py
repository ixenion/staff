#work with excel files (libreoffice etc.)
#openpyxl library is used

import openpyxl as xl
from openpyxl.chart import BarChart, Reference



wb = xl.load_workbook('example.xlsx')
sheet = wb['Sheet1']#to access first page of file
cell = sheet['a1']#returns cell A1 containment
cell = sheet.cell(1, 1)#returns exact same cell
sheet.max_row#returns max (last) row number

for row in range(1, sheet.max_row + 1):
    corrected_price = sheet.cell(row, 3) * 0.9#to get aces to the partiular cell and to reduce price for 10%
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

#creating a chart (graphic)
values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('example2.xlsx')

##########################################################################
##same but with functions##

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']#to access first page of file
    cell = sheet['a1']#returns cell A1 containment
    cell = sheet.cell(1, 1)#returns exact same cell
    sheet.max_row#returns max (last) row number

    for row in range(1, sheet.max_row + 1):
        corrected_price = sheet.cell(row, 3) * 0.9#to get aces to the partiular cell and to reduce price for 10%
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

#creating a chart (graphic)
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

